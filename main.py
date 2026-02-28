# in this version price api dont work

from telethon import TelegramClient, events, Button
from openpyxl import Workbook, load_workbook
from khayyam import JalaliDatetime
import aiohttp
import asyncio
import uuid
import os
import re

# set developer and manager
DEVELOPER = 1477966103 # @amirhosin282
MANAGER = 7474464502 # sepehr electronic

# importing data from env
with open("./env/api_id.txt", "r", encoding="utf-8-sig") as api_id_file:
    API_ID = int(api_id_file.read().strip()) # get api id

with open("./env/api_hash.txt", "r") as api_hash_file:
    API_HASH = api_hash_file.read().strip() # get api hash

with open("./env/token.txt", "r") as token_file:
    TOKEN = token_file.read().strip() # get bot token

with open("./env/admin.txt", "r", encoding="utf-8-sig") as admin_chat_id:
    admins = admin_chat_id.read().strip().split(",") # get admin id

with open("./data/answerd_questions.txt", "r", encoding="utf-8-sig") as asqe:
    answerd_qestion_list = asqe.read().split(",") # read answeard question after start

with open("./env/api_key.txt", "r", encoding="utf-8-sig") as api_key:
    api_key = api_key.read().strip()

# set answeard questions
question_key = {}
answerd_qestion = set()
for messages_key in answerd_qestion_list:
    answerd_qestion.add(messages_key)
    question_key[messages_key] = "0"
print("answeard questions loaded")
    

# load excel file in set dir
user_data = set() # chat ids set var
FILE = "./data/users.xlsx"

# create excel file if not created before
if not os.path.exists(FILE) :
    wb = Workbook()
    ws = wb.active
    ws.title = "users_data"
    # write titels in file
    ws.append([
        "chat_id",
        "user_name",
        "first_name",
        "last_name"
        
    ])
    wb.save(FILE)
    wb.close()


# cheng price format 
def format_price(price: int) -> str:
    return f"{price: ,}".replace(",", ".")
    

# return date time
def date_ret():
    now = JalaliDatetime.now()
    time = now.strftime("%Y/%m/%d - %H:%M:%S")
    return time


# get prices 
api_url = "https://brsapi.ir/Api/Market/Gold_Currency.php"
params_ = {"key": api_key}

headers_ = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json",
    "Referer": "https://brsapi.ir/"
}

server_error = True

async def get_api(u, p, h):
    global server_error
    async with aiohttp.ClientSession(headers=h) as session:
        while True:
            try:
                async with session.get(u, params=p) as res:
                    if res.status == 200:
                        data = await res.json()
                        print("data updated")
                    else:
                        text = await res.text()
                        print("bad status:", res.status, "|", text[:200])

            except aiohttp.ClientConnectorError as e:
                print("Connection error:", e)
                server_error = True
            except Exception as e:
                print("API error:", e)
                server_error = True

            await asyncio.sleep(3600)  # ever 1 houer

# get chat id from excel file into set var
wb =  load_workbook(FILE)
ws = wb.active
for i in ws.iter_rows(min_row= 2, values_only= True):
    user_data.add(i[0])

print("loaded users chat ids")

# check data in excel and if not exist add id to file and set 
def add_data(set_var, file_dir, dict):
    if dict["chat_id"] not in set_var:
        wb = load_workbook(file_dir)
        ws = wb.active
        # writ to file
        ws.append([
            dict["chat_id"],
            dict["user_name"],
            dict["first_name"],
            dict["last_name"]
        ])
        
        set_var.add(dict["chat_id"])
        wb.save(file_dir)
        wb.close()

# creating app
client = TelegramClient(
    "./data/sepehr_bot",
    api_id= API_ID,
    api_hash= API_HASH
)

# main func
async def main(token, app, admin_id, date):
     # start app
    await app.start(bot_token= token)
       
    # send start data to admin
    for i in admin_id:
        await app.send_message(int(i), f"bot started in {date} succses")
    
    # update prices
    asyncio.create_task(get_api(api_url, params_, headers_))

    # keeping run the bot
    await app.run_until_disconnected()
    
# start message
@client.on(events.NewMessage(pattern= r"/start"))
async def start(event):
    sender = await event.get_sender()
    data ={
        "chat_id" : event.chat_id,
        "user_name" : sender.username,
        "first_name" : sender.first_name,
        "last_name" : sender.last_name
    }
    # use add data func to write to excel file
    add_data(user_data, FILE, data)
    print(f"start request from : {data['chat_id']}")
    
    # set services keyboared button
    global keyBoard_services_button
    keyBoard_services_button = client.build_reply_markup([
        [
            Button.text("❓ پرسش سؤال", resize=True),
            Button.text("🏪 لیست خدمات حضوری", resize=True)
        ],
        
        [
            Button.text("🧰 سامانه‌های کاربردی", resize=True),
            Button.text("💰 قیمت لحظه‌ای طلا و ارزها", resize=True),
        ], # sort buttons two by two
        
        [Button.text("تمامی خدمات 🛠️", resize=True)],
        [Button.text("پنل ادمین 👨🏻‍💻", resize=True) if str(event.chat_id) in admins else Button.clear()] # set this button when requester is admin
    ])
    
    await client.send_message(entity= event.chat_id, message= f"سلام {data['first_name']} عزیز، \n به ربات خدمات کامپیوتری سپهر خوش آمدید \n برای مشاهده خدمات از دستور /services استفاده کنید.", buttons= keyBoard_services_button)

# admin panel
@ client.on(events.NewMessage(pattern=r"^(/admin|پنل ادمین 👨🏻‍💻)$"))
async def adminPanel(event):
    # set buttom
    butt = client.build_reply_markup([
        [Button.inline("📤 ارسال پیام", data="send_to_all")],
        [Button.inline("➕ افزودن ادمین", data="add_admin")],
        [Button.inline("📂 دریافت دیتابیس و لیست ادمین‌ها", data="get_db")],
        [Button.inline("❌ حذف ادمین", data="remove_admin")],
        [Button.inline("💬 پاسخ به سوالات کاربران", data="reply_users")]
    ])

    if str(event.chat_id) in admins:
        print(f"Confirmed admin login with chat ID: {event.chat_id} detected")
        sender = await event.get_sender()
        # set text
        text = f"""
        سلام مدیر عزیز ({sender.first_name}) 👋
        به پنل مدیریت ربات خوش آمدید

        💡 گزینه‌ها را لمس کنید:

        📤 ارسال پیام به همه کاربران
        ➕ افزودن ادمین
        ❌ حذف ادمین
        📂 دریافت دیتابیس و لیست ادمین‌ها
        💬 پاسخ به سوالات کاربران
        """

        await client.send_message(entity= event.chat_id, message= text, buttons= butt)
    else:
        print(f"An unverified request with chat ID: {event.chat_id} was detected to log in to the admin panel")


# send servises
@client.on(events.NewMessage(pattern= r"(/services|تمامی خدمات 🛠️)$"))
async def services(event):
    print(f"services reqest from : {event.chat_id}")
    
    # set butons
    butt = client.build_reply_markup([
        [Button.inline("🧰 سامانه‌های کاربردی", data="usefull_sites"),
         Button.url("📸 اینستاگرام ما", "https://www.instagram.com/sepehr_._electronic/")],

        [Button.url("🚗 ثبت‌نام ایران‌خودرو", "https://ikcosales.ir/"),
         Button.inline("🏪 لیست خدمات حضوری", data="services_list")],

        [Button.inline("❓ پرسش سؤال", data="ask"),
         Button.inline("💰 قیمت لحظه‌ای طلا و ارزها", data="price")]
    ])


    # set text for view
    text = """
    📌 <b>خدمات ربات سپهر الکترونیک</b>

    🧰 <b>سامانه‌های کاربردی</b>
    مجموعه‌ای از لینک‌ها و ابزارهای آنلاین پرکاربرد

    🚗 <b>ثبت‌نام ایران‌خودرو</b>
    ورود مستقیم به سامانه فروش

    🏪 <b>لیست خدمات حضوری</b>
    مشاهده خدمات و تعمیرات

    ❓ <b>پرسش سؤال</b>
    ارتباط مستقیم با پشتیبانی

    💰 <b>قیمت لحظه‌ای طلا و رمزارزها</b>
    بررسی قیمت‌های روز بازار

    📸 <b>اینستاگرام ما</b>
    نمونه‌کارها و اطلاعیه‌ها
    """

    # send message
    await client.send_message(entity=event.chat_id, message=text, buttons= butt, parse_mode= "html")
 
 
# send message to all 
@client.on(events.CallbackQuery(data="send_to_all"))
async def sendToAll(event):
    
    # set back button
    cancel_keyboard = [
        [Button.text("❌ لغو", resize=True, single_use=True)]
    ]
    
    if str(event.chat_id) in admins:
        # print log and get response
        print(f"admin {event.chat_id} want to send to all a message")
        
        async with client.conversation(event.chat_id, timeout=500) as conv:
            await conv.send_message("متن ارسالی خود را وارد کنید: ", buttons=cancel_keyboard)
            response = await conv.get_response()
            response = response.text
            
            # cancel buttone
            if response == "❌ لغو":
                await client.send_message(event.chat_id, message= "عملیات با موفقیت لغو شد \n برای بازگشت به پنل ادمین از دستور /admin استفاده کنید.", buttons=keyBoard_services_button)
                return
            
            # send message to all ids in user_data set var
            for ids in user_data:
                await client.send_message(entity=ids, message=response, buttons=None)
            return

# add new admin
@client.on(events.CallbackQuery(data="add_admin"))
async def add_admin(event):
    if str(event.chat_id) in admins:
        # set back button
        cancel_keyboard = [
            [Button.text("❌ لغو", resize=True, single_use=True)]
        ]

        if str(event.chat_id) in admins:
            print(f"admin {event.chat_id} want to add a new admin")

            #start conversation
            async with client.conversation(event.chat_id, timeout=500) as conv:
                await conv.send_message("آیدی عددی ادمین جدید را وارد کنید: ", buttons=cancel_keyboard)
                response = await conv.get_response()
                new_admin = response.text.strip()

                # cancle
                if new_admin == "❌ لغو":
                    await client.send_message(event.chat_id, message= "عملیات با موفقیت لغو شد \n برای بازگشت به پنل ادمین از دستور /admin استفاده کنید.", buttons=keyBoard_services_button)
                    return

                # checing admin
                if new_admin in admins:
                    await client.send_message(event.chat_id, "⚠️ این کاربر از قبل ادمین است", buttons=keyBoard_services_button)
                    return

                # check inital message
                if not new_admin.isdigit():
                    await client.send_message(event.chat_id, "❌ آیدی باید عددی باشد", buttons=keyBoard_services_button)
                    return

                else: # add to admins
                    with open("./env/admin.txt", "a", encoding="utf-8-sig") as admin_chat_id:
                        admin_chat_id.write(f",{str(new_admin).strip()}")
                    admins.append(new_admin)
                    print(f"admin {event.chat_id} added {new_admin} to admins")
                    await client.send_message(int(new_admin), message="شما به ادمین ارتقاع یافتید\n برای مشاهده دسترسی های خود از دستور /admin استفاده کنید")
                    await client.send_message(int(new_admin), message= "برای بازدهی بهتر ربات،‌ توصیه میشه یه بار مجدد ربات رو /start بکنید!!", buttons= keyBoard_services_button)
                    await client.send_message(event.chat_id, message= "ادمین جدید اضافه شد", buttons=keyBoard_services_button)
                    return
    else:
        return
    

# send data base to admin
@client.on(events.CallbackQuery(data="get_db"))
async def get_db(event):
    if str(event.chat_id) in admins:
        print(f"admin {event.chat_id} want to get databeses")
        await client.send_file(event.chat_id, ["./data/users.xlsx", "./data/sepehr_bot.session", "./data/answerd_questions.txt", "./env/admin.txt", "./env/token.txt", "./env/api_hash.txt", "./env/api_id.txt"])
        return
    else:
        return


# removing an admin
@client.on(events.CallbackQuery(data="remove_admin"))
async def remove_admin(event):
    if str(event.chat_id) in admins:
        # set back button
        cancel_keyboard = [
            [Button.text("❌ لغو", resize=True, single_use=True)]
        ]

        if str(event.chat_id) in admins:
            # print log
            print(f"admin {event.chat_id} want to remove an admin")

            # start conversation
            async with client.conversation(event.chat_id, timeout=500) as conv:
                await conv.send_message("آیدی عددی ادمینی که قصد بر حذفشان دارید را وارد کنید: ", buttons=cancel_keyboard)
                response = await conv.get_response()
                r_chat_id = response.text

            # checking text
            if r_chat_id == "❌ لغو":
                await client.send_message(event.chat_id, message="عملیات با موفقیت لغو شد", buttons=keyBoard_services_button)
                return

            elif int(r_chat_id) == DEVELOPER or int(r_chat_id) == MANAGER:
                await client.send_message(event.chat_id, message="شما نمیتوانید مدیریت یا توسعه دهنده ربات را حذف کنید", buttons=keyBoard_services_button)
                return

            elif r_chat_id not in admins:
                await client.send_message(event.chat_id, message=f"چت آیدی ارسالی شما {r_chat_id} در لیست مدیران موجود نمیباشد", buttons=keyBoard_services_button)
                return

            elif r_chat_id == str(event.chat_id):
                await client.send_message(event.chat_id, message= " شما نمیتوانید دسترسی ادمین را از خود بگیرید", buttons=keyBoard_services_button)
                return

            else: # removing admin
                with open("./env/admin.txt", "r", encoding="utf-8") as admin_chat_id:
                    admin_chat_id = admin_chat_id.read().replace(f",{r_chat_id}", "")

                with open("./env/admin.txt", "w", encoding="utf-8") as new_admins:
                    new_admins.write(admin_chat_id)

                admins.remove(r_chat_id)
                print(f"admin {event.chat_id} removed {r_chat_id} admin")

                text = f"""
                ادمین مورد نظر حذف شد \n ادمین حذف شده : {r_chat_id} \n ادمین های موجود : {admins} \n مدریت : {MANAGER} \n توسعه دهنده : {DEVELOPER}
                """
                await client.send_message(event.chat_id, message= text, buttons= keyBoard_services_button)
                return
    else:
        return


# get question and send to admins
@client.on(events.CallbackQuery(data="ask"))
@client.on(events.NewMessage(pattern="❓ پرسش سؤال"))
async def send_to_admin(event):
    
    global keyBoard_services_button

    print(f"user {event.chat_id} want ask a question")
    # set cancel button
    cancel_keyboard = [
        [Button.text("❌ لغو", resize=True, single_use=True)]
    ]
    # start conversation
    sender = await event.get_sender() # get sender data
    text = """
    💬 سوالت رو برام بنویس و من می‌فرستم برای ادمین‌ها.
    📨 بعد از اینکه ادمین جواب داد، جوابش رو برات می‌فرستم.
    """
    async with client.conversation(event.chat_id, timeout=500) as conv:
        await conv.send_message(text, buttons=cancel_keyboard)
        response = await conv.get_response()
        question = response.text
        # check cancle
        if question == "❌ لغو":
            await client.send_message(event.chat_id, message="عملیات با موفقیت لغو شد", buttons=keyBoard_services_button)
            return
    text = f"""
    🧾 سوال جدید از کاربر:
    👤 کاربر: @{sender.username}
    🧑‍💼 نام: {sender.first_name}
    📝 نام خانوادگی: {sender.last_name if sender.last_name else '—'}
    🆔 آیدی: {event.chat_id}
    📅 تاریخ: {date_ret()}
    ❓ سوال:
    {question}
    برای پاسخ دادن روی دکمه‌ی زیر بزن:
    """
    # answer button
    question_uid = uuid.uuid4().hex[:8]
    question_key [question_uid] = event.chat_id
    data_str = f"answer:{question_uid}"
    data_byt= data_str.encode()
    answer = [
        [Button.inline("✉️ پاسخ", data=data_byt)]
    ]
    for admin in admins:
        await client.send_message(int(admin), message= text, buttons= answer)
    await client.send_message(event.chat_id, message= "پیام با موفقیت برای ادمین ها ارسال شد ✅", buttons=keyBoard_services_button)
    print(f"user {event.chat_id} send question for admins")
    

# send answer
@client.on(events.CallbackQuery(data=re.compile(b"^answer:")))
async def answer(event):
    if str(event.chat_id) in admins:
    
        # set back button
        cancel_keyboard = [
            [Button.text("❌ لغو", resize=True, single_use=True)]
        ]

        # get message id and sender chat id
        message = await event.get_message()
        message_id = message.id

        data = event.data.decode().split(":")
        message_uid = data[1]
        sender_id = question_key[message_uid]


        if str(message_uid) in answerd_qestion:
            await client.send_message(event.chat_id, message=f"به این سوال قبلا پاسخ داده شده❌", buttons=keyBoard_services_button)
            return
        else:
            # print log
            print(f"admin {event.chat_id} want to answer to {message_uid} - {sender_id} question")

            # start conversation
            async with client.conversation(event.chat_id, timeout=500) as conv:
                await conv.send_message("🖊 لطفاً جواب رو بنویس و ارسال کن:", buttons=cancel_keyboard)
                response = await conv.get_response()
                answer = response.text


            # set sending text
            send_text=f"""
            📬 پاسخ شما از طرف ادمین دریافت شد:

            {answer}
            """

            # check cancel
            if answer == "❌ لغو":
                await client.send_message(event.chat_id, message="عملیات با موفقیت لغو شد", buttons=keyBoard_services_button)
                return
            else:
                # send answer to user
                await client.send_message(sender_id, message=send_text, buttons=keyBoard_services_button)

                # delete question from admin and set question key
                await client.delete_messages(event.chat_id, message_ids=message_id)

                with open("./data/answerd_questions.txt", "a", encoding="utf-8-sig") as asqe:
                    asqe.write(f"{message_uid},")
                answerd_qestion.add(message_uid)

                # print log
                print(f"admin {event.chat_id} answerd to {message_uid} - {sender_id} question")
                await client.send_message(event.chat_id, message= "پاسخ با موفقیت ارسال شد ✅", buttons= keyBoard_services_button)
    else:
        return

@client.on(events.CallbackQuery(data="reply_users"))
async def reply_to_user_ask(event):
    if str(event.chat_id) in admins:
        await client.send_message(event.chat_id, message= "سوال های ارسالی کاربران مستقیما به تمامی ادمین ها ارسال میشه و همراه با متن سوال یه گزینه برای پاسخگویی ادمین ها میاد، برای حواب دادن به سوالا فقط کافیه از دکمه ی مربوط به مون سوال استفاده کنی") 
        return
    else: 
        return


# price of gold and digital market
@client.on(events.CallbackQuery(data="price"))
@client.on(events.NewMessage(pattern="💰 قیمت لحظه‌ای طلا و ارزها"))
async def prise(event):
    butt = client.build_reply_markup([
        Button.inline("طلا", data = "gold"),
        Button.inline("ارز های دیجیتال", data = "market"),
        Button.inline("دلار و ارز های دیگر", data = "real")
    ])

    
    # send message to user
    await client.send_message(
        entity=event.chat_id, 
        message= """
            📊 جهت مشاهده قیمت‌ها، یکی از گزینه‌های زیر را انتخاب نمایید:
            🟡 طلا
            🪙 ارزهای دیجیتال
            💲 دلار و ارز های دیگر
            """, 
            buttons= butt)

# gold price 
@client.on(events.CallbackQuery(data="gold"))
async def gold_price(event):
    print(f"user {event.chat_id} requested for gold price")
    # check server
    if server_error == False :
        pass # the web service server is has error in connection so this parts (gold - market - real) dont work on this version
    
    else :
        # send error to user
        await client.send_message(entity= event.chat_id, message="⚠️خطای اتصال به وب سرویس،‌ لطفا دقایی دیگر مجددا تلاش کنید⚠️", buttons= keyBoard_services_button)
        return

# market prie 
@client.on(events.CallbackQuery(data="market"))
async def gold_price(event):
    print(f"user {event.chat_id} requested for market price")
    # check server
    if server_error == False :
        pass # the web service server is has error in connection so this parts (gold - market - real) dont work on this version
    
    else :
        # send error to user
        await client.send_message(entity= event.chat_id, message="⚠️خطای اتصال به وب سرویس،‌ لطفا دقایی دیگر مجددا تلاش کنید⚠️", buttons= keyBoard_services_button)
        return

@client.on(events.CallbackQuery(data="real"))
async def gold_price(event):
    print(f"user {event.chat_id} requested for real price")
    # check server
    if server_error == False :
        pass # the web service server is has error in connection so this parts (gold - market - real) dont work on this version
    
    else :
        # send error to user
        await client.send_message(entity= event.chat_id, message="⚠️خطای اتصال به وب سرویس،‌ لطفا دقایی دیگر مجددا تلاش کنید⚠️", buttons= keyBoard_services_button)
        return


# send usefull websites adress
@client.on(events.CallbackQuery(data= "usefull_sites"))
@client.on(events.NewMessage(pattern="🧰 سامانه‌های کاربردی"))
async def usefull_sites(event):
    await client.send_file(entity= event.chat_id, file= "./data/usefull_sites.pdf")
    await client.send_message(entity= event.chat_id, message= "نکته : اکثر لینک های داخل لیس قابلیت انتقال مستقیم به سامانه دارن،فقط کافیه روشون کلیک کنید!", buttons= keyBoard_services_button)
    print(f"user {event.chat_id} get the usefull sites")
    return


# send services 
@client.on(events.CallbackQuery(data = "services_list"))
@client.on(events.NewMessage(pattern="🏪 لیست خدمات حضوری"))
async def services_in(event):
    text = """
        📌 لیست خدمات کافی‌نت و خدمات کامپیوتری سپهر

        🖥 خدمات کامپیوتری و نرم‌افزاری
        • نصب ویندوز (7 / 10 / 11)
        • نصب لینوکس (Ubuntu / Debian / Arch)
        • نصب درایور و نرم‌افزارهای عمومی و تخصصی
        • نصب آنتی‌ویروس و تأمین امنیت سیستم
        • پارتیشن‌بندی و فرمت هارد
        • رفع کندی سیستم و بهینه‌سازی
        • ویروس‌یابی و پاکسازی کامل
        • رفع مشکل بوت و بالا نیامدن ویندوز
        • ریکاوری اطلاعات
        • بازیابی فایل‌های حذف‌شده
        • بکاپ‌گیری و انتقال اطلاعات
        • ریست پسورد ویندوز
        • ارتقاء سیستم و نصب SSD

        🌐 خدمات اینترنتی و سامانه‌های دولتی
        • ثبت‌نام دانشگاه، مدارس و آزمون‌ها
        • ثبت‌نام خودرو و موتور
        • خدمات ثنا، سجام، سهام عدالت
        • ثبت‌نام یارانه و کارت سوخت
        • پرداخت قبوض و جریمه
        • دریافت فیش حقوقی و گواهی‌ها
        • احراز هویت آنلاین

        🖨 خدمات چاپ و تکثیر
        • پرینت سیاه‌وسفید و رنگی (A4 / A5 / A3)
        • چاپ پشت و رو
        • کپی سیاه‌وسفید و رنگی
        • اسکن مدارک
        • پرینت عکس
        • تایپ فارسی و انگلیسی

        📝 امور اداری و حقوقی
        • تنظیم و ثبت دادخواست
        • ثبت شرکت و تغییرات
        • اظهارنامه مالیاتی
        • نقل و انتقال خودرو
        • خدمات بیمه تأمین اجتماعی

        💾 خدمات فایل و حافظه
        • رایت CD / DVD
        • انتقال اطلاعات فلش و هارد
        • فرمت و تعمیر فلش
        • تبدیل و کم‌حجم‌سازی فایل‌ها

        📱 خدمات موبایل
        • نصب نرم‌افزار
        • انتقال اطلاعات گوشی
        • ساخت جیمیل
        • رفع مشکلات نرم‌افزاری اندروید

        🎯 خدمات ویژه
        • نصب کامل سیستم (پکیج ویندوز + درایور + آفیس + آنتی‌ویروس + بهینه‌سازی)
        • راه‌اندازی شبکه کوچک
        • تنظیم مودم و اینترنت
        • طراحی رزومه حرفه‌ای
        • مشاوره خرید سیستم
    """
    await client.send_message(event.chat_id, message= text, buttons= keyBoard_services_button)
    return


# start bot    
if __name__ == "__main__":
    asyncio.run(main(TOKEN, client, admins, date_ret()))